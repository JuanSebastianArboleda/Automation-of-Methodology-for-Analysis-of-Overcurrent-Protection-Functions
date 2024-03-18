# -*- coding: utf-8 -*-
"""
Created on Mon Dec 21 18:01:28 2020

@author: Juan Sebastian Arboleda Arroyave
Versión: 0
"""

################################################# IMPORTAR LIBRERIAS Y LECTURA INFORMACIÓN ##################################################
import powerfactory as pf  # Importa la librería de Digsilent /Archivos de programa x86/ Digsilent/Python
import math  # Falta SEL DIR
import sys
import pandas as pd
import os
import time
from collections import OrderedDict

app=pf.GetApplication()                 # Inicia la aplicación
app.ClearOutputWindow()
script=app.GetCurrentScript()
app.EchoOff()

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


Address=script.Address
xlsx=script.xlsx
Change=script.Change
iSpgf=script.iSpgf  
iBpgf=script.iBpgf  
iMethod=script.Method    
Terminal=script.Terminal 
Barrido=script.Barrido
Diagnostico=script.Diagnostico

Reboot=[]
app.PrintInfo('Copyright Juan Sebastian Arboleda y asociados - Todos los derechos reservados ©')
app.PrintPlain(' ')
if Change=='Y' or Change=='y':
    app.PrintInfo('Sí se realizarán cambios en los ajustes presentados a continuación')
    Reboot.append(Address)
    Reboot.append(xlsx)
    Reboot.append('Y')
    Reboot.append(iSpgf)
    Reboot.append(iBpgf)
    Reboot.append(iMethod)
    Reboot.append(Terminal)
    Reboot.append(Barrido)
    Reboot.append(Diagnostico)
    script.IntExpr=(Reboot)
elif Change!='Y' or Change!='y':
    app.PrintInfo('No se realizarán cambios en los ajustes presentados a continuación')
    app.PrintError('No se realizarán cambios en los ajustes presentados a continuación')
iSpgf=int(script.iSpgf)
iBpgf=int(script.iBpgf)
iMethod=int(script.Method)
Terminal=int(script.Terminal)
Barrido=int(script.Barrido)
Diagnostico=int(script.Diagnostico)

Address=str(Address)+'\\'+str(xlsx)+'.xlsx'
oShc=script.GetContents('Shc')[0]
MatDis=script.GetContents('IntMat_Distancias')[0]
MatRes=script.GetContents('IntMat_Resistencias')[0]
wb=load_workbook(filename=Address, data_only=True)
Project=app.GetActiveProject()
Reles=script.GetContents('Librería de Relés.IntFolder')[0]
Eq_med=Reles.GetContents('Equipos de Medida')[0]
PTs_p=Eq_med.GetContents('PT HMV.TypVt')[0]
PTs_s=Eq_med.GetContents('SecundarioPT HMV.TypVtsec')[0]
CT=Eq_med.GetContents('CT HMV.TypCt')[0]
sElms=script.GetContents('Element Model')[0].All()
lElms=script.GetContents('Element Falla')[0].All()
sCs=script.GetContents('Cases')[0].GetAll('IntCase')
Relays=script.GetContents('Relays')[0]
Relays.Clear()

if Terminal==0:
    Term='i'
else:
    Term='j'

start=time.time()

if sCs==[]:
    sCs.append(app.GetActiveStudyCase())
################################################# Transformadores de medida #################################################

def sub_Crear_TM(iRele):
    for Elm in sElms:
        if Elm.loc_name == ws.cell(row=10+iRele, column=2).value:
            if Elm.GetClassName()=='ElmLne':
                bus1=Elm.GetAttribute('bus1')
                bus2=Elm.GetAttribute('bus2')
                sElmsLn.append(Elm)
            elif Elm.GetClassName()=='ElmBranch':
                bus1s=Elm.GetAttribute('p_conn0')
                bus2s=Elm.GetAttribute('p_conn1')
                bus1=bus1s.GetAttribute('bus1')
                bus2=bus2s.GetAttribute('bus2')
                sElmsLn.append(Elm)
            elif Elm.GetClassName()=='ElmCoup':
                bus1=Elm.GetAttribute('bus1')
                bus2=Elm.GetAttribute('bus2')
                sElmsTer.append(Elm)
            elif Elm.GetClassName()=='ElmTr3':
                bus1=Elm.GetAttribute('bushv')
                bus2=Elm.GetAttribute('busmv')
                bus3=Elm.GetAttribute('buslv')
                sElmsTer.append(Elm)
            elif Elm.GetClassName()=='ElmTr2':
                bus1=Elm.GetAttribute('bushv')
                bus3=Elm.GetAttribute('buslv')
                sElmsTer.append(Elm)
            elif Elm.GetClassName()=='ElmSym':
                bus1=Elm.GetAttribute('bus1')
                sElmsTer.append(Elm)
            Term = ws.cell(row=10+iRele, column=3).value
            ##### CT #####
            lCT_p = ws.cell(row=10+iRele, column=12).value
            lCT_s = ws.cell(row=10+iRele, column=13).value
            if lCT_p == None or lCT_s == None:
                app.PrintError('Verifique los valores de corriente del CT de la fila '+str(10+iRele)+', columnas L y M')
                sys.exit()
            
            if Term=='i':
                if bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus1.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
            elif Term=='j':
                if bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus2.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
            
            elif Term=='HV':
                if bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus1.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
                
            elif Term=='MV':
                if bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus2.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus2.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
                
            elif Term=='LV':
                if bus3.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus3.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus3.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus3.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
                
            elif Term=='GEN':
                if bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt'):
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                else:
                    bus1.CreateObject('StaCt', 'CT_'+str(lCT_p)+'-'+str(lCT_s))
                    contents=bus1.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
                sub_CT_Type(CT, contents, lCT_p, lCT_s)
            else:
                app.PrintError('No se encontró un terminal valido en la fila '+str(10+iRele)+' de la columna C')
                sys.exit()
            ##### PT #####
            lPT_p=ws.cell(row=10+iRele, column=10).value
            lPT_s=ws.cell(row=10+iRele, column=11).value
            if lPT_p == None or lPT_s == None:
                app.PrintError('Verifique los valores de tensión del PT de la fila '+str(10+iRele)+', columnas J y K')
                sys.exit()
            if Term=='i':
                if bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus1.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            elif Term=='j':        
                if bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus2.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            elif Term=='HV':        
                if bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus1.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            elif Term=='MV':        
                if bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus2.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus2.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            elif Term=='LV':        
                if bus3.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus3.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus3.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus3.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            elif Term=='GEN':        
                if bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt'):
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                else:
                    bus1.CreateObject('StaVt', 'PT_'+str(lPT_p)+'-'+str(lPT_s))
                    contents=bus1.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
                sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s)
            else:
                app.PrintError('No se encontró un terminal valido en la fila '+str(10+iRele)+' de la columna C')
                sys.exit()

def sub_CT_Type(CT, contents, lCT_p, lCT_s):
    contents.typ_id = CT
    contents.ptapset=lCT_p
    contents.stapset=lCT_s
def sub_PT_Type(PTs_p, PTs_s, contents, lPT_p, lPT_s):
    contents.typ_id=PTs_p
    contents.staptyp=PTs_s
    contents.ptapset=lPT_p
    contents.stapset=lPT_s

def sub_Crear_Reles(iRele):
    iE=0
    for Elm in sElms:
        if Elm.loc_name == ws.cell(row=10+iRele, column=2).value:
            iElms.append(lElms[iE].loc_name)
            Term = ws.cell(row=10+iRele, column=3).value
            sExt = ws.cell(row=10+iRele, column=4).value
            sRel = ws.cell(row=10+iRele, column=7).value
            if sExt == None:
                app.PrintError('No se encontró un nombre valido de extremo en la fila '+str(10+iRele)+' de la columna D')
                sys.exit()
            elif sRel == None:
                app.PrintError('No se encontró un nombre valido de relé en la fila '+str(10+iRele)+' de la columna G')
                sys.exit()
            Rel_name = str(sExt)+'_'+str(sRel)
            Rele_name = ws.cell(row=10+iRele, column=8).value
            
            if Elm.GetClassName()=='ElmLne':
                if Term == 'i':
                    bus=Elm.GetAttribute('bus1')
                elif Term == 'j':
                    bus=Elm.GetAttribute('bus2')
            elif Elm.GetClassName()=='ElmBranch':
                if Term == 'i':
                    bus1s=Elm.GetAttribute('p_conn0')
                    bus=bus1s.GetAttribute('bus1')
                elif Term == 'j':
                    bus2s=Elm.GetAttribute('p_conn1')
                    bus=bus2s.GetAttribute('bus2')
            elif Elm.GetClassName()=='ElmCoup':
                if Term == 'i':
                    bus=Elm.GetAttribute('bus1')
                elif Term == 'j':
                    bus=Elm.GetAttribute('bus2')
            
            elif Elm.GetClassName()=='ElmTr3':
                if Term == 'HV':
                    bus=Elm.GetAttribute('bushv')
                elif Term == 'MV':
                    bus=Elm.GetAttribute('busmv')
                elif Term == 'LV':
                    bus=Elm.GetAttribute('buslv')
            
            elif Elm.GetClassName()=='ElmTr2':
                if Term == 'HV':
                    bus=Elm.GetAttribute('bushv')
                elif Term == 'LV':
                    bus=Elm.GetAttribute('buslv')
            
            elif Elm.GetClassName()=='ElmSym':
                if Term == 'GEN':
                    bus=Elm.GetAttribute('bus1')
            else:
                app.PrintError('Error, con elemento '+str(Elm.loc_name)+' verifique celda '+str(10+iRele)+' columna C')
            
            lCT_p = ws.cell(row=10+iRele, column=12).value
            lCT_s = ws.cell(row=10+iRele, column=13).value
            lPT_p=ws.cell(row=10+iRele, column=10).value
            lPT_s=ws.cell(row=10+iRele, column=11).value
            CT=bus.GetContents('CT_'+str(lCT_p)+'-'+str(lCT_s)+'.StaCt')[0]
            PT=bus.GetContents('PT_'+str(lPT_p)+'-'+str(lPT_s)+'.StaVt')[0]
            cnt=0
            if ws.cell(row=10+iRele, column=16).value == 'Forward' or ws.cell(row=10+iRele, column=20).value == 'Forward' or ws.cell(row=10+iRele, column=16).value == 'Reverse' or ws.cell(row=10+iRele, column=20).value == 'Reverse':
                if ws.cell(row=10+iRele, column=23).value == 'None' or ws.cell(row=10+iRele, column=28).value == 'None':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-51N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-51N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_67-51N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-51N.ElmRelay')[0]
                elif ws.cell(row=10+iRele, column=23).value == 'Forward' or ws.cell(row=10+iRele, column=28).value == 'Forward' or ws.cell(row=10+iRele, column=23).value == 'Reverse' or ws.cell(row=10+iRele, column=28).value == 'Reverse':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-67N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-67N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_67-67N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67-67N.ElmRelay')[0]
                else:
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_67')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67.ElmRelay')[0]
                    
            elif ws.cell(row=10+iRele, column=16).value == 'None' or ws.cell(row=10+iRele, column=20).value == 'None':
                if ws.cell(row=10+iRele, column=23).value == 'None' or ws.cell(row=10+iRele, column=28).value == 'None':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-51N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-51N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_51-51N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-51N.ElmRelay')[0]
                elif ws.cell(row=10+iRele, column=23).value == 'Forward' or ws.cell(row=10+iRele, column=28).value == 'Forward' or ws.cell(row=10+iRele, column=23).value == 'Reverse' or ws.cell(row=10+iRele, column=28).value == 'Reverse':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-67N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-67N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_51-67N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51-67N.ElmRelay')[0]
                else:
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_51')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51.ElmRelay')[0]
            
            else:
                if ws.cell(row=10+iRele, column=23).value == 'None' or ws.cell(row=10+iRele, column=28).value == 'None':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_51N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_51N.ElmRelay')[0]
                elif ws.cell(row=10+iRele, column=23).value == 'Forward' or ws.cell(row=10+iRele, column=28).value == 'Forward' or ws.cell(row=10+iRele, column=23).value == 'Reverse' or ws.cell(row=10+iRele, column=28).value == 'Reverse':
                    if bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67N.ElmRelay'):
                        Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67N.ElmRelay')[0]
                    else:
                        bus.CreateObject('ElmRelay', str(iRele)+'_'+str(Rel_name)+'_67N')    
                    Add_IED=bus.GetContents(str(iRele)+'_'+str(Rel_name)+'_67N.ElmRelay')[0]
                else:
                    app.PrintError('No se creó el relé de la fila '+str(10+iRele)+' debido a que estan apagadas todas sus funciones')
                    cnt=1
                    break
            Relays.AddRef(Add_IED)
            if cnt!=1:
                if Rele_name == 'REL 670':
                    Rele_Type = Reles.GetContents('ABB.IntFolder')[0]
                    IED_typ = Rele_Type.GetContents('REL 670.TypRelay')[0]
                elif Rele_name == 'GE D60':
                    Rele_Type = Reles.GetContents('GE.IntFolder')[0]
                    IED_typ = Rele_Type.GetContents('UR D60.TypRelay')[0]
                elif Rele_name == 'SEL 421':
                    Rele_Type = Reles.GetContents('SEL.IntFolder')[0]
                    if CT.stapset == 1:
                        IED_typ = Rele_Type.GetContents('SEL 421-1A.TypRelay')[0]
                    elif CT.stapset == 5:
                        IED_typ = Rele_Type.GetContents('SEL 421-5A.TypRelay')[0]
                elif Rele_name == '7SA6' :
                    IED_Folder = Reles.GetContents('SIEMENS.IntFolder')[0]
                    if CT.stapset == 1:
                        IED_typ = IED_Folder.GetContents('7SA6 1A.TypRelay')[0]
                    elif CT.stapset == 5:
                        IED_typ = IED_Folder.GetContents('7SA6 5A.TypRelay')[0]
                Add_IED.typ_id = IED_typ
                IED_List.append(Add_IED.loc_name)
                sub_Ingresar_ajustes(Rele_name, Add_IED, CT, PT, Elm, Term, lCT_p, IED_typ)
        iE+=1

def sub_Ingresar_ajustes(Rele_name, Add_IED, CT, PT, Elm, Term, lCT_p, IED_typ):
    if Rele_name == 'REL 670':
        app.PrintPlain('Se encontraron los siguientes cambios en el relé '+str(Add_IED))
        sub_51_REL670(Add_IED, CT, PT, lCT_p, IED_typ)
    elif Rele_name == '7SA6':
        app.PrintPlain('Se encontraron los siguientes cambios en el relé '+str(Add_IED))
        sub_51_7SA6(Add_IED, CT, PT, lCT_p, IED_typ)
    elif Rele_name == 'SEL 421':
        app.PrintPlain('Se encontraron los siguientes cambios en el relé '+str(Add_IED))
        Sub_Rutina_SEL(Add_IED, CT, PT, lCT_p, IED_typ, Elm)
    elif Rele_name == 'GE D60':
        app.PrintPlain('Se encontraron los siguientes cambios en el relé '+str(Add_IED))
        Sub_Rutina_GE(Add_IED, CT, PT, lCT_p, IED_typ)

def sub_51_REL670(Add_IED, CT, PT, lCT_p, IED_typ):
    Measure=Add_IED.GetContents('Measure.RelMeasure')[0]
    RMeasure=Add_IED.GetContents('Remote Measure.RelMeasure')[0]
    DMeasure=Add_IED.GetContents('Delta Measure.RelMeasure')[0]
    ArmMeasure=Add_IED.GetContents('2nd harm Measure.RelMeasure')[0]
    Measure.Inom=CT.stapset
    Measure.Unom=PT.stapset
    RMeasure.Inom=CT.stapset
    DMeasure.Inom=CT.stapset
    DMeasure.Unom=PT.stapset
    ArmMeasure.Inom=CT.stapset
    
    PDIF_87=Add_IED.GetContents('PDIF 87.ElmRelay')[0]
    PIOC=Add_IED.GetContents('PIOC.ElmRelay')[0]
    PTOF_PTUF_PFRC_81=Add_IED.GetContents('PTOF PTUF PFRC 81.ElmRelay')[0]
    PTUV_27_PTOV_5959N=Add_IED.GetContents('PTUV 27 PTOV 59/59N.ElmRelay')[0]
    RPSB_78=Add_IED.GetContents('RPSB 78.ElmRelay')[0]
    SC_DirZ=Add_IED.GetContents('SC Dir-Z.ElmRelay')[0]
    ZMCPDIS=Add_IED.GetContents('ZMCPDIS.ElmRelay')[0]
    ZMHPDIS=Add_IED.GetContents('ZMHPDIS.ElmRelay')[0]
    ZMQPDIS=Add_IED.GetContents('ZMQPDIS.ElmRelay')[0]
    DirZ=Add_IED.GetContents('Dir-Z.RelDisdir')[0]
    PHS_Load_Area=Add_IED.GetContents('PHS Load Area.RelDisloadenc')[0]
    PDIF_87.outserv=1
    PIOC.outserv=1
    PTOF_PTUF_PFRC_81.outserv=1
    PTUV_27_PTOV_5959N.outserv=1
    RPSB_78.outserv=1
    SC_DirZ.outserv=1
    ZMCPDIS.outserv=1
    ZMHPDIS.outserv=1
    ZMQPDIS.outserv=1
    DirZ.outserv=1
    PHS_Load_Area.outserv=1
    
    disel=Add_IED.pdiselm
    disel[0]=CT
    disel[1]=CT
    disel[2]=PT
    Add_IED.pdiselm=disel
    
    PTOC_51_67=Add_IED.GetContents('PTOC 51_67.ElmRelay')[0]
    PTOC_51N67N=Add_IED.GetContents('PTOC 51N67N.ElmRelay')[0]

    if ws.cell(row=10+iRele, column=14).value != '%':
        app.PrintError('Para relés ABB no se recibe otro tipo de datos diferente al %, revisar fila '+str(10+iRele)+', columna N')        
        sys.exit()
    if ws.cell(row=10+iRele, column=16).value == 'Off' and ws.cell(row=10+iRele, column=20).value == 'Off':
        PTOC_51_67.outserv=1
    else:
        PTOC_51_67.outserv=0
    if ws.cell(row=10+iRele, column=23).value == 'Off' and ws.cell(row=10+iRele, column=28).value == 'Off':
        PTOC_51N67N.outserv=1
    else:
        PTOC_51N67N.outserv=0
        
    # 67-51 / 67N-51N
    for FT in range(2):
        if FT==0:
            sVar67=['idir','Ipset','pcharac','Tpset']
            PTOC_51_67_1=PTOC_51_67.GetContents('PTOC 51_67 1.RelToc')[0]
            DirPTOC=PTOC_51_67.GetContents('Dir PTOC.RelDir')[0]
            DirPTOC.outserv=0
            if ws.cell(row=10+iRele, column=16).value != 'Off':
                PTOC_51_67_1.outserv=0
                PTOC_51_67_1.ModFrame=1
            else:
                PTOC_51_67_1.outserv=1
        elif FT==1:
            sVar67=['idir','Ipset','pcharac','Tpset','Tadder']
            PTOC_51_67_1=PTOC_51N67N.GetContents('PTOC 51N_67N 1.RelToc')[0]
            DirPTOCN=PTOC_51N67N.GetContents('Dir PTOC N.RelDir')[0]
            DirPTOCN.outserv=0
            if ws.cell(row=10+iRele, column=23).value != 'Off':
                PTOC_51_67_1.outserv=0
                PTOC_51_67_1.ModFrame=1
            else:
                PTOC_51_67_1.outserv=1
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=16+FT*7, max_col=19+FT*7+FT, min_row=10+iRele, max_row=10+iRele):
                    for cell in row:
                        if ws.cell(row=10+iRele, column=16+FT*7).value != 'Off':
                            if cell.value == None:
                                app.PrintError('Celda vacía '+str(cell))
                                sys.exit()
                            else:
                                Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(PTOC_51_67_1.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                PTOC_51_67_1.SetAttribute(sVar67[i],1)
                            else:
                                PTOC_51_67_1.SetAttribute(sVar67[i],2)
                    elif str(PTOC_51_67_1.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                PTOC_51_67_1.SetAttribute(sVar67[i],0)
                            else:
                                PTOC_51_67_1.SetAttribute(sVar67[i],2)
                    elif str(PTOC_51_67_1.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                PTOC_51_67_1.SetAttribute(sVar67[i],0)
                            else:
                                PTOC_51_67_1.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipset':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                    Aprim = Ibase*iPkup
                    iPkpn = Aprim/lCT_p
                    if abs(float(PTOC_51_67_1.GetAttribute(sVar67[i])*lCT_p) - float(iPkup*Ibase)) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(PTOC_51_67_1.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkup*Ibase))+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            PTOC_51_67_1.SetAttribute(sVar67[i],iPkpn)
                elif sVar67[i]=='pcharac':
                    Valor=sub_Curva_ABB(Data[i])
                    if str(PTOC_51_67_1.GetAttribute(sVar67[i]).loc_name) != str(Valor):
                        app.PrintPlain('Diferencia en valores de Digsilent '+str(PTOC_51_67_1.GetAttribute(sVar67[i]).loc_name)+' y Excel '+ str(Valor)+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            TCCs=IED_typ.GetContents('TCCs.IntFolder')[0]
                            Curves=TCCs.GetContents('*.TypChatoc')
                            for Curve in Curves:
                                if Curve.loc_name == Valor: 
                                    PTOC_51_67_1.SetAttribute(sVar67[i],Curve)
                elif sVar67[i]=='Tpset':
                    if abs(float(PTOC_51_67_1.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(PTOC_51_67_1.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            PTOC_51_67_1.SetAttribute(sVar67[i],Data[i])
                elif sVar67[i]=='Tadder':
                    if abs(float(PTOC_51_67_1.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(PTOC_51_67_1.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_1))
                        if Change == 'Y' or Change == 'y':
                            PTOC_51_67_1.SetAttribute(sVar67[i],Data[i])
    # 50-50N
    for FT in range(2):
        if FT==0:
            sVar67=['idir','Ipset','Tpset']
            PTOC_51_67_2=PTOC_51_67.GetContents('PTOC 51_67 2.RelToc')[0]
            if ws.cell(row=10+iRele, column=20).value != 'Off':
                PTOC_51_67_2.outserv=0
                PTOC_51_67_2.ModFrame=1
            else:
                PTOC_51_67_2.outserv=1
        elif FT==1:
            sVar67=['idir','Ipset','Tpset']
            PTOC_51_67_2=PTOC_51N67N.GetContents('PTOC 51N_67N 2.RelToc')[0]
            if ws.cell(row=10+iRele, column=28).value != 'Off':
                PTOC_51_67_2.outserv=0
                PTOC_51_67_2.ModFrame=1
            else:
                PTOC_51_67_2.outserv=1
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=20+FT*8, max_col=22+FT*8, min_row=10+iRele, max_row=10+iRele):
                    for cell in row:
                        if ws.cell(row=10+iRele, column=20+FT*8).value != 'Off':
                            if cell.value == None:
                                app.PrintError('Celda vacía '+str(cell))
                                sys.exit()
                            else:
                                Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(PTOC_51_67_2.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_2))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                PTOC_51_67_2.SetAttribute(sVar67[i],1)
                            else:
                                PTOC_51_67_2.SetAttribute(sVar67[i],2)
                    elif str(PTOC_51_67_2.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_2))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                PTOC_51_67_2.SetAttribute(sVar67[i],0)
                            else:
                                PTOC_51_67_2.SetAttribute(sVar67[i],2)
                    elif str(PTOC_51_67_2.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_2))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                PTOC_51_67_2.SetAttribute(sVar67[i],0)
                            else:
                                PTOC_51_67_2.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipset':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    iPkup = ws.cell(row=10+iRele, column=21+FT*8).value
                    Aprim = Ibase*iPkup
                    iPkpn = Aprim/lCT_p
                    if abs(float(PTOC_51_67_2.GetAttribute(sVar67[i])*lCT_p) - float(iPkup*Ibase)) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(PTOC_51_67_2.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkup*Ibase))+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_2))
                        if Change == 'Y' or Change == 'y':
                            PTOC_51_67_2.SetAttribute(sVar67[i],iPkpn)
                elif sVar67[i]=='Tpset':
                    if abs(float(PTOC_51_67_2.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(PTOC_51_67_2.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(PTOC_51_67_2))
                        if Change == 'Y' or Change == 'y':
                            PTOC_51_67_2.SetAttribute(sVar67[i],Data[i])
                Valor='Definite time TCC'
                if str(PTOC_51_67_2.GetAttribute('pcharac').loc_name) != str(Valor):
                    app.PrintPlain('Diferencia en valores de Digsilent '+str(PTOC_51_67_2.GetAttribute('pcharac').loc_name)+' y Excel '+ str(Valor)+' para parametro pcharac en el bloque '+str(PTOC_51_67_2))
                    if Change == 'Y' or Change == 'y':
                        TCCs=IED_typ.GetContents('TCCs.IntFolder')[0]
                        Curves=TCCs.GetContents('*.TypChatoc')
                        for Curve in Curves:
                            if Curve.loc_name == Valor: 
                                PTOC_51_67_2.SetAttribute('pcharac',Curve)

def sub_51_7SA6(Add_IED, CT, PT, lCT_p, IED_typ):
    Measurement=Add_IED.GetContents('Measurement.RelMeasure')[0]
    DMeasure=Add_IED.GetContents('Measurement Delta.RelMeasure')[0]
    MeasSeq=Add_IED.GetContents('Measurement Seq.RelMeasure')[0]
    MeasI0mut=Add_IED.GetContents('Measurement I4.RelMeasure')[0]
    MeasFreq=Add_IED.GetContents('Measurement Freq.RelFmeas')[0]
    Measurement.Inom=CT.stapset
    Measurement.Unom=PT.stapset
    DMeasure.Inom=CT.stapset
    DMeasure.Unom=PT.stapset
    MeasSeq.Inom=CT.stapset
    MeasSeq.Unom=PT.stapset
    MeasI0mut.Inom=CT.stapset
    MeasFreq.Unom=PT.stapset
    
    Distance_Polygonal=Add_IED.GetContents('Distance Polygonal.ElmRelay')[0]
    Frequency_50Hz=Add_IED.GetContents('Frequency (50Hz).ElmRelay')[0]
    Frequency_60Hz=Add_IED.GetContents('Frequency (60Hz).ElmRelay')[0]
    Thermal_Overload=Add_IED.GetContents('Thermal Overload.RelToc')[0]
    Voltage=Add_IED.GetContents('Voltage.ElmRelay')[0]
    Distance_Polygonal.outserv=1
    Frequency_50Hz.outserv=1
    Frequency_60Hz.outserv=1
    Voltage.outserv=1
    Thermal_Overload.outserv=1
    
    Overcurrent=Add_IED.GetContents('Overcurrent (50/51/67)')[0]
    try:
        Earthfault=Add_IED.GetContents('Overcurrent (50N/51N/67N).ElmRelay')[0]
    except:
        if ws.cell(row=10+iRele, column=13).value == 1:
            Earthfault=Add_IED.GetContents('Overcurrent (50N/51N/67N).ElmRelay')[0]   # 1 A
        elif ws.cell(row=10+iRele, column=13).value == 5:
            Earthfault=Add_IED.GetContents('Earthfault (51N/67N).ElmRelay')[0]   # 5 A
    
    Ip67DIR=Overcurrent.GetContents('Dir OC.RelDir')[0]
    Ip67NDIR=Earthfault.GetContents('Directional (I0-U0).RelDir')[0]
    Ip67NDIR1=Earthfault.GetContents('Directional (S).RelDir')[0]
    Ip67NDIR2=Earthfault.GetContents('Directional (I2-U2).RelDir')[0]
    Ip67DIR.outserv=0
    Ip67NDIR.outserv=0
    Ip67NDIR1.outserv=0
    Ip67NDIR2.outserv=0
    disel=Add_IED.pdiselm
    disel[0]=CT
    disel[1]=PT
    disel[2]=CT
    Add_IED.pdiselm=disel

    Unidad=ws.cell(row=10+iRele, column=14).value
    if Unidad != 'A prim' and Unidad != 'A sec':
        app.PrintError('Los relés SIEMENS solo reciben sus unidades en A prim o A sec, favor revisar fila '+str(10+iRele)+', columna N')        
        sys.exit()
    if ws.cell(row=10+iRele, column=16).value == 'Off' and ws.cell(row=10+iRele, column=20).value == 'Off':
        Overcurrent.outserv=1
    else:
        Overcurrent.outserv=0
    if ws.cell(row=10+iRele, column=23).value == 'Off' and ws.cell(row=10+iRele, column=28).value == 'Off':
        Earthfault.outserv=1
    else:
        Earthfault.outserv=0

    # 67-51 / 67N-51N
    for FT in range(2):
        if FT==0:
            sVar67=['idir','Ipsetr','pcharac','Tpset']
            try:
                Ip67TOC=Overcurrent.GetContents('Ip 67 TOC.RelToc')[0]
            except:
                Ip67TOC=Overcurrent.GetContents('Ip>.RelToc')[0]
            if ws.cell(row=10+iRele, column=16).value != 'Off':
                Ip67TOC.outserv=0
                Ip67TOC.ModFrame=1
            else:
                Ip67TOC.outserv=1
        elif FT==1:
            sVar67=['idir','Ipsetr','pcharac','Tpset','Tadder']
            Ip67TOC=Earthfault.GetContents('3I0p.RelToc')[0]
            if ws.cell(row=10+iRele, column=23).value != 'Off':
                Ip67TOC.outserv=0
                Ip67TOC.ModFrame=1
            else:
                Ip67TOC.outserv=1
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=16+FT*7, max_col=19+FT*7+FT, min_row=10+iRele, max_row=10+iRele):
                    for cell in row:
                        if ws.cell(row=10+iRele, column=16+FT*7).value != 'Off':
                            if cell.value == None:
                                app.PrintError('Celda vacía '+str(cell))
                                sys.exit()
                            else:
                                Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Ip67TOC.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                Ip67TOC.SetAttribute(sVar67[i],1)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipsetr':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        Asec = iPkup*Icts/Ictp
                        iPkpn = Asec
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                            if Change == 'Y' or Change == 'y':
                                Ip67TOC.SetAttribute(sVar67[i],iPkpn)
                    else:
                        iPkpn = ws.cell(row=10+iRele, column=17+FT*7).value
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(iPkpn))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                            if Change == 'Y' or Change == 'y':
                                Ip67TOC.SetAttribute(sVar67[i],iPkpn)
                            
                elif sVar67[i]=='pcharac':
                    Valor=sub_Curva_SIEMENS(Data[i])
                    if str(Ip67TOC.GetAttribute(sVar67[i]).loc_name) != str(Valor):
                        app.PrintPlain('Diferencia en valores de Digsilent '+str(Ip67TOC.GetAttribute(sVar67[i]).loc_name)+' y Excel '+ str(Valor)+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Characteristics=IED_typ.GetContents('Characteristics.IntFolder')[0]
                            Curves=Characteristics.GetContents('*.TypChatoc')
                            for Curve in Curves:
                                if Curve.loc_name == Valor: 
                                    Ip67TOC.SetAttribute(sVar67[i],Curve)
                elif sVar67[i]=='Tpset':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])
                elif sVar67[i]=='Tadder':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])
    
    # 50-50N
    for T50 in range(2):
        if T50==0:
            sVar67=['idir','Ipsetr','Tset']
            if ws.cell(row=10+iRele, column=13).value == 1:
                Iph=Overcurrent.GetContents('Iph>.RelIoc')[0]  # 1 A
            elif ws.cell(row=10+iRele, column=13).value == 5:
                Iph=Overcurrent.GetContents('Iph>.RelIoc')[0]  # 5 A
            if ws.cell(row=10+iRele, column=20).value != 'Off':
                Iph.outserv=0
                Iph.ModFrame=1
            else:
                Iph.outserv=1
        elif T50==1:
            sVar67=['idir','Ipsetr','Tset']
            Iph=Earthfault.GetContents('3I0>.RelIoc')[0]
            if ws.cell(row=10+iRele, column=28).value != 'Off':
                Iph.outserv=0
                Iph.ModFrame=1
            else:
                Iph.outserv=1
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=20+T50*8, max_col=22+T50*8, min_row=10+iRele, max_row=10+iRele):
                    for cell in row:
                        if ws.cell(row=10+iRele, column=20+T50*8).value != 'Off':
                            if cell.value == None:
                                app.PrintError('Celda vacía '+str(cell))
                                sys.exit()
                            else:
                                Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Iph.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                Iph.SetAttribute(sVar67[i],1)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipsetr':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=21+T50*8).value
                        Asec = iPkup*Icts/Ictp
                        iPkpn = Asec
                        if abs(float(Iph.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    else:
                        iPkpn = ws.cell(row=10+iRele, column=21+T50*8).value
                        if abs(float(Iph.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(iPkpn))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    if Change == 'Y' or Change == 'y':
                        Iph.SetAttribute(sVar67[i],iPkpn)
                elif sVar67[i]=='Tset':
                    if abs(float(Iph.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            Iph.SetAttribute(sVar67[i],Data[i])
            
def Sub_Rutina_SEL(Add_IED, CT, PT, lCT_p, IED_typ, Elm):
    Measurement=Add_IED.GetContents('M-Iab/Ibc/Ica.RelMeasure')[0]
    MeasMI=Add_IED.GetContents('M-I/U.RelMeasure')[0]
    MeasRMS=Add_IED.GetContents('Meas RMS seq.RelMeasure')[0]
    Measurement.Inom=CT.stapset
    MeasMI.Unom=PT.stapset
    MeasMI.Unom=PT.stapset
    MeasRMS.Inom=CT.stapset
    MeasRMS.Unom=PT.stapset
    L27=Add_IED.GetContents('27L.RelUlim')[0]
    L59=Add_IED.GetContents('59L.RelUlim')[0]
    Load_Encroachment=Add_IED.GetContents('Load Encroachment.RelDisloadenc')[0]
    L27.outserv=1
    L59.outserv=1
    Load_Encroachment.outserv=1
    for i in range(5):
        ZMP=Add_IED.GetContents('Z'+str(i+1)+'MP.RelDismho')[0]
        ZMG=Add_IED.GetContents('Z'+str(i+1)+'MG.RelDismho')[0]
        ZMQP=Add_IED.GetContents('Z'+str(i+1)+'QP.RelDispoly')[0]
        ZMQG=Add_IED.GetContents('Z'+str(i+1)+'QG.RelDispoly')[0]
        ZPD=Add_IED.GetContents('Z'+str(i+1)+'PD.RelTimer')[0]
        ZGD=Add_IED.GetContents('Z'+str(i+1)+'GD.RelTimer')[0]
        ZMP.outserv=1
        ZMG.outserv=1
        ZMQP.outserv=1
        ZMQG.outserv=1
        ZPD.outserv=1
        ZGD.outserv=1
    
    OOS=Add_IED.GetContents('Out of step.ElmRelay')[0]
    B51S1N=Add_IED.GetContents('51S1N.RelToc')[0]
    B51S1I1=Add_IED.GetContents('51S1P.RelToc')[0]
    B51S1P=Add_IED.GetContents('51S1I1.RelToc')[0]
    B51S2I1=Add_IED.GetContents('51S2I1.RelToc')[0]
    B51S2N=Add_IED.GetContents('51S2N.RelToc')[0]
    B51S2P=Add_IED.GetContents('51S2P.RelToc')[0]
    B51S3I1=Add_IED.GetContents('51S3I1.RelToc')[0]
    B51S3N=Add_IED.GetContents('51S3N.RelToc')[0]
    B51S3P=Add_IED.GetContents('51S3P.RelToc')[0]
    B51S1Q=Add_IED.GetContents('51S1Q.RelToc')[0]
    B51S2Q=Add_IED.GetContents('51S2Q.RelToc')[0]
    B51S3Q=Add_IED.GetContents('51S3Q.RelToc')[0]
    
    B50N1=Add_IED.GetContents('50/67N1.RelIoc')[0]
    
    B50N2=Add_IED.GetContents('50/67N2.RelIoc')[0]
    B50N3=Add_IED.GetContents('50/67N3.RelIoc')[0]
    B50N4=Add_IED.GetContents('50/67N4.RelIoc')[0]
    B50P1=Add_IED.GetContents('50/67P1.RelIoc')[0]
    B50P2=Add_IED.GetContents('50/67P2.RelIoc')[0]
    B50P3=Add_IED.GetContents('50/67P3.RelIoc')[0]
    B50P4=Add_IED.GetContents('50/67P4.RelIoc')[0]
    B46Q1=Add_IED.GetContents('50/67Q1.RelIoc')[0]
    B46Q2=Add_IED.GetContents('50/67Q2.RelIoc')[0]
    B46Q3=Add_IED.GetContents('50/67Q3.RelIoc')[0]
    B46Q4=Add_IED.GetContents('50/67Q4.RelIoc')[0]
    
    OOS.outserv=1
    B51S1P.outserv=1
    B51S2I1.outserv=1
    B51S2N.outserv=1
    B51S2P.outserv=1
    B51S3I1.outserv=1
    B51S3N.outserv=1
    B51S3P.outserv=1
    B51S1Q.outserv=1
    B51S2Q.outserv=1
    B51S3Q.outserv=1
    
    B50N2.outserv=1
    B50N3.outserv=1
    B50N4.outserv=1
    B50P2.outserv=1
    B50P3.outserv=1
    B50P4.outserv=1
    B46Q1.outserv=1
    B46Q2.outserv=1
    B46Q3.outserv=1
    B46Q4.outserv=1
    
    disel=Add_IED.pdiselm
    disel[0]=CT
    disel[1]=PT
    Add_IED.pdiselm=disel
    
    ### SEL DIR ###
    if Elm.GetClassName()=='ElmLne' or Elm.GetClassName()=='ElmBranch':
        Z1=Elm.GetAttribute('e:Z1')
        R0=Elm.GetAttribute('e:R0')
        if R0==0:
            R0=0.00001
        X0=Elm.GetAttribute('e:X0')
        Ilinea=Elm.GetAttribute('e:Inom')*1000
        Vl=Elm.GetAttribute('e:Unom')*1000
        Bc=Elm.GetAttribute('e:B0')/1000000
        Vg=Vl/math.sqrt(3)
        Z0=math.sqrt(R0**2+X0**2)
        Z0ANGrad=math.atan(X0/R0)
        PTR=float(PT.ptapset)/float(PT.stapset)
        CTR=float(CT.ptapset)/float(CT.stapset)
        K=CTR/PTR
        Z1MAG=K*Z1
        Z1ANG=Elm.GetAttribute('e:phiz1')
        Z0MAG=K*Z0
        Z0ANG=math.degrees(Z0ANGrad)
        InomLne=Ilinea*(1/CTR)
        SelDir=Add_IED.GetContents('Sel Dir.RelSeldir')[0]
        SelDir.s50QF=0.12*InomLne
        SelDir.s50QR=0.08*InomLne
        SelDir.Z2F=0.5*Z1MAG
        SelDir.Z2R=(0.5*Z1MAG)+(0.5/InomLne)
        SelDir.a2=0.1
        SelDir.k2=0.2
        SelDir.Zm=Z1MAG
        SelDir.phi=Z1ANG
        SelDir.s50GFP=0.12*InomLne
        SelDir.s50GRP=0.08*InomLne
        SelDir.Z0F=0.5*Z0MAG
        SelDir.Z0R=(0.5*Z0MAG)+(0.5/InomLne)
        SelDir.s50LP=Vg*Bc
        SelDir.a0=0.1
        SelDir.Z0=Z0MAG
        SelDir.phi0=Z0ANG
    
    Unidad=ws.cell(row=10+iRele, column=14).value
    if Unidad != 'A prim' and Unidad != 'A sec' and Unidad != 'p.u':
        app.PrintError('Los relés SEL solo reciben sus unidades en A prim o A sec, favor revisar fila '+str(10+iRele)+', columna N')        
        sys.exit()
        
    if ws.cell(row=10+iRele, column=16).value == 'Off':
        B51S1I1.outserv=1
    else:
        B51S1I1.outserv=0
        
    if ws.cell(row=10+iRele, column=23).value == 'Off':
        B51S1N.outserv=1
    else:
        B51S1N.outserv=0
        
    if ws.cell(row=10+iRele, column=20).value == 'Off':
        B50P1.outserv=1
    else:
        B50P1.outserv=0
    if ws.cell(row=10+iRele, column=28).value == 'Off':
        B50N1.outserv=1
    else:
        B50N1.outserv=0
        
    # 67-51 / 67N-51N
    for FT in range(2):
        if FT==0:
            sVar67=['idir','Ipsetr','pcharac','Tpset']
            Ip67TOC=B51S1I1
        elif FT==1:
            sVar67=['idir','Ipsetr','pcharac','Tpset','Tadder']
            Ip67TOC=B51S1N
            Ip67TOC.ModFrame=1
        
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=16+FT*7, max_col=19+FT*7+FT, min_row=10+iRele, max_row=10+iRele):
            for cell in row:
                if ws.cell(row=10+iRele, column=16+FT*7).value != 'Off':
                    if cell.value == None:
                        app.PrintError('Celda vacía '+str(cell))
                        sys.exit()
                    else:
                        Data.append(cell.value)
                        
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Ip67TOC.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                
                                Ip67TOC.SetAttribute(sVar67[i],1)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipsetr':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        Pu = iPkup/Ictp
                        iPkpn = Pu
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    elif Unidad == 'A sec':
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        Pu = iPkup/Icts
                        iPkpn = Pu
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])*Icts))+' y Excel ''{0:.2f}'.format(float(iPkpn*Icts))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    else:
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        iPkpn = iPkup
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(iPkpn)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(iPkpn))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    if Change == 'Y' or Change == 'y':
                        Ip67TOC.SetAttribute(sVar67[i],iPkpn)
                            
                elif sVar67[i]=='pcharac':
                    Valor=sub_Curva_SEL(Data[i])
                    if str(Ip67TOC.GetAttribute(sVar67[i]).loc_name) != str(Valor):
                        app.PrintPlain('Diferencia en valores de Digsilent '+str(Ip67TOC.GetAttribute(sVar67[i]).loc_name)+' y Excel '+ str(Valor)+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Characteristics=IED_typ.GetContents('Characteristics.IntFolder')[0]
                            Curves=Characteristics.GetContents('*.TypChatoc')
                            for Curve in Curves:
                                if Curve.loc_name == Valor: 
                                    Ip67TOC.SetAttribute(sVar67[i],Curve)
                elif sVar67[i]=='Tpset':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])
                elif sVar67[i]=='Tadder':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])

    # 50-50N
    for T50 in range(2):
        if T50==0:
            sVar67=['idir','Ipsetr','cTset']
            Iph=B50P1
            Iph.ModFrame=1
        elif T50==1:
            sVar67=['idir','Ipsetr','cTset']
            Iph=B50N1
            Iph.ModFrame=1
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=20+T50*8, max_col=22+T50*8, min_row=10+iRele, max_row=10+iRele):
            for cell in row:
                if ws.cell(row=10+iRele, column=20+T50*8).value != 'Off':
                    if cell.value == None:
                        app.PrintError('Celda vacía '+str(cell))
                        sys.exit()
                    else:
                        Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Iph.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                Iph.SetAttribute(sVar67[i],1)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipsetr':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=21+T50*8).value
                        Asec = iPkup*Icts/Ictp
                        iPkpn = Asec
                        if abs(float(Iph.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    elif Unidad == 'A sec':
                        iPkpn = ws.cell(row=10+iRele, column=21+T50*8).value
                        if abs(float(Iph.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(iPkpn))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    else:
                        pu = ws.cell(row=10+iRele, column=21+T50*8).value
                        iPkpn = pu/Icts
                        if abs(float(Iph.GetAttribute(sVar67[i])*Ictp/Icts) - float(pu)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])*Ictp/Icts))+' y Excel ''{0:.2f}'.format(float(pu))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    if Change == 'Y' or Change == 'y':
                        Iph.SetAttribute(sVar67[i],iPkpn)
                elif sVar67[i]=='cTset':
                    if abs((float(Iph.GetAttribute(sVar67[i]))/60) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i]))/60)+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro cptotime en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            Iph.SetAttribute('cTset',Data[i]*60)

def Sub_Rutina_GE(Add_IED, CT, PT, lCT_p, IED_typ):
    MeasMutual=Add_IED.GetContents('Meas Mutual.RelMeasure')[0]
    MeasNeutral=Add_IED.GetContents('Meas Neutral I.RelMeasure')[0]
    Measdelta=Add_IED.GetContents('Meas delta.RelMeasure')[0]
    Measurement=Add_IED.GetContents('Measurement.RelMeasure')[0]
    MeasSeq=Add_IED.GetContents('Measurement Seq.RelMeasure')[0]
    MeasMutual.Inom=CT.stapset
    MeasNeutral.Unom=PT.stapset
    MeasNeutral.Inom=CT.stapset
    Measdelta.Inom=CT.stapset
    Measdelta.Unom=PT.stapset
    Measurement.Inom=CT.stapset
    Measurement.Unom=PT.stapset
    MeasSeq.Inom=CT.stapset
    MeasSeq.Unom=PT.stapset
        
    Ground_Distance_elms=Add_IED.GetContents('Ground Distance elements (F21).ElmRelay')[0]
    Phase_Distance_elms=Add_IED.GetContents('Phase Distance elements (F21).ElmRelay')[0]
    Voltage_elms=Add_IED.GetContents('Voltage elements (F27 - F59).ElmRelay')[0]
    Ground_Distance_elms.outserv=1
    Phase_Distance_elms.outserv=1
    Voltage_elms.outserv=1
    Overcurrent=Add_IED.GetContents('Overcurrent elements (F50 - F51 - F46).ElmRelay')[0]
    
    disel=Add_IED.pdiselm
    disel[0]=CT
    disel[1]=PT
    disel[2]=CT
    disel[3]=CT
    Add_IED.pdiselm=disel

    Unidad=ws.cell(row=10+iRele, column=14).value
    if Unidad != 'A prim' and Unidad != 'A sec':
        app.PrintError('Los relés GE solo reciben sus unidades en A prim o A sec, favor revisar fila '+str(10+iRele)+', columna N')        
        sys.exit()
    if ws.cell(row=10+iRele, column=16).value == 'Off' and ws.cell(row=10+iRele, column=20).value == 'Off' and ws.cell(row=10+iRele, column=23).value == 'Off' and ws.cell(row=10+iRele, column=28).value == 'Off':
        Overcurrent.outserv=1
    else:
        Overcurrent.outserv=0
    
    NegSeqToc1=Overcurrent.GetContents('Negative sequence Toc 1.RelToc')[0]
    NegSeqToc2=Overcurrent.GetContents('Negative sequence Toc 2.RelToc')[0]
    NegSeqIoc1=Overcurrent.GetContents('Negative sequence Ioc 1.RelIoc')[0]
    NegSeqIoc2=Overcurrent.GetContents('Negative sequence Ioc 2.RelIoc')[0]
    NegSeqToc1.outserv=1
    NegSeqToc2.outserv=1
    NegSeqIoc1.outserv=1
    NegSeqIoc2.outserv=1
    
    Unidad=ws.cell(row=10+iRele, column=14).value
    if Unidad != 'A prim' and Unidad != 'A sec' and Unidad != 'p.u':
        app.PrintError('Los relés GE solo reciben sus unidades en A prim, A sec o p.u; favor revisar fila '+str(10+iRele)+', columna N')        
        sys.exit()
        
    PhaseToc1=Overcurrent.GetContents('Phase Toc 1.RelToc')[0]
    GroundToc1=Overcurrent.GetContents('Ground Toc 1.RelToc')[0]
    NeutralToc1=Overcurrent.GetContents('Neutral Toc 1.RelToc')[0]
    PhaseIoc1=Overcurrent.GetContents('Phase Ioc 1.RelIoc')[0]
    GroundIoc1=Overcurrent.GetContents('Ground Ioc 1.RelIoc')[0]
    NeutralIoc1=Overcurrent.GetContents('Neutral Ioc 1.RelIoc')[0]
    PhaseToc2=Overcurrent.GetContents('Phase Toc 2.RelToc')[0]
    GroundToc2=Overcurrent.GetContents('Ground Toc 2.RelToc')[0]
    NeutralToc2=Overcurrent.GetContents('Neutral Toc 2.RelToc')[0]
    PhaseIoc2=Overcurrent.GetContents('Phase Ioc 2.RelIoc')[0]    
    GroundIoc2=Overcurrent.GetContents('Ground Ioc 2.RelIoc')[0]    
    NeutralIoc2=Overcurrent.GetContents('Neutral Ioc 2.RelIoc')[0]
    PhaseToc2.outserv=1
    GroundToc2.outserv=1
    NeutralToc2.outserv=1
    PhaseIoc2.outserv=1
    GroundIoc2.outserv=1
    NeutralIoc2.outserv=1
    
    DirPhase1=Overcurrent.GetContents('DirPhase 1.RelDir')[0]
    DirPhase2=Overcurrent.GetContents('DirPhase 2.RelDir')[0]
    DirNeutral1=Overcurrent.GetContents('DirNeutral 1.RelDir')[0]
    DirNeutral2=Overcurrent.GetContents('DirNeutral 2.RelDir')[0]
    DirPhase1.outserv=0
    DirPhase2.outserv=0
    DirNeutral1.outserv=0
    DirNeutral2.outserv=0
    
    if ws.cell(row=10+iRele, column=16).value == 'Off':
        PhaseToc1.outserv=1
    else:
        PhaseToc1.outserv=0
        PhaseToc1.ModFrame=1
    if ws.cell(row=10+iRele, column=23).value == 'Off':
        GroundToc1.outserv=1
        NeutralToc1.outserv=1
    else:
        GroundToc1.outserv=1
        NeutralToc1.outserv=0
#        GroundToc1.ModFrame=1
        NeutralToc1.ModFrame=1
        
    if ws.cell(row=10+iRele, column=20).value == 'Off':
        PhaseIoc1.outserv=1
    else:
        PhaseIoc1.outserv=0
        PhaseIoc1.ModFrame=0
    
    if ws.cell(row=10+iRele, column=28).value == 'Off':
        GroundIoc1.outserv=1
        NeutralIoc1.outserv=1
    else:
        GroundIoc1.outserv=1
        NeutralIoc1.outserv=0
        GroundIoc1.ModFrame=0
        NeutralIoc1.ModFrame=0
        
    # 67-51 / 67N-51N
    for FT in range(2):
        if FT==0:
            sVar67=['idir','Ipset','pcharac','Tpset']
            Ip67TOC=PhaseToc1

        elif FT==1:
            sVar67=['idir','Ipset','pcharac','Tpset','Tadder']
            Ip67TOC=GroundToc1
            Ip67TOCN=NeutralToc1

        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=16+FT*7, max_col=19+FT*7+FT, min_row=10+iRele, max_row=10+iRele):
                    for cell in row:
                        if ws.cell(row=10+iRele, column=16+FT*7).value != 'Off':
                            if cell.value == None:
                                app.PrintError('Celda vacía '+str(cell))
                                sys.exit()
                            else:
                                Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Ip67TOC.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                Ip67TOC.SetAttribute(sVar67[i],1)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],1)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],2)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],2)
                    elif str(Ip67TOC.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Ip67TOC.SetAttribute(sVar67[i],0)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],0)
                            else:
                                Ip67TOC.SetAttribute(sVar67[i],1)
                                if FT==1:
                                    Ip67TOCN.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipset':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        pu = iPkup/Ictp
                        iPkpn = pu
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    elif Unidad == 'A sec':
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        pu = iPkup/Icts
                        iPkpn = pu
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])*Icts) - float(iPkpn*Icts)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])*Icts))+' y Excel ''{0:.2f}'.format(float(iPkpn*Icts))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    else:
                        iPkup = ws.cell(row=10+iRele, column=17+FT*7).value
                        iPkpn = iPkup
                        if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(iPkpn)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(iPkpn))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                    if Change == 'Y' or Change == 'y':
                        Ip67TOC.SetAttribute(sVar67[i],iPkpn)
                        if FT==1:
                            Ip67TOCN.SetAttribute(sVar67[i],iPkpn)
                            
                elif sVar67[i]=='pcharac':
                    Valor=sub_Curva_GE(Data[i])
                    if str(Ip67TOC.GetAttribute(sVar67[i]).loc_name) != str(Valor):
                        app.PrintPlain('Diferencia en valores de Digsilent '+str(Ip67TOC.GetAttribute(sVar67[i]).loc_name)+' y Excel '+ str(Valor)+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Characteristics=IED_typ.GetContents('Characteristics.IntFolder')[0]
                            Curves=Characteristics.GetContents('*.TypChatoc')
                            for Curve in Curves:
                                if Curve.loc_name == Valor: 
                                    Ip67TOC.SetAttribute(sVar67[i],Curve)
                                    if FT==1:
                                        Ip67TOCN.SetAttribute(sVar67[i],Curve)
                elif sVar67[i]=='Tpset':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])
                            if FT==1:
                                Ip67TOCN.SetAttribute(sVar67[i],Data[i])
                elif sVar67[i]=='Tadder':
                    if abs(float(Ip67TOC.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Ip67TOC.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Ip67TOC))
                        if Change == 'Y' or Change == 'y':
                            Ip67TOC.SetAttribute(sVar67[i],Data[i])
                            if FT==1:
                                Ip67TOCN.SetAttribute(sVar67[i],Data[i])
    # 50-50N
    for T50 in range(2):
        if T50==0:
            sVar67=['idir','Ipset','Tset']
            Iph=PhaseIoc1
        elif T50==1:
            sVar67=['idir','Ipset','Tset']
            Iph=GroundIoc1
            Iph1=NeutralIoc1
            
        Data=[]     # Lectura de datos
        for row in ws.iter_rows(min_col=20+T50*8, max_col=22+T50*8, min_row=10+iRele, max_row=10+iRele):
            for cell in row:
                if ws.cell(row=10+iRele, column=20+T50*8).value != 'Off':
                    if cell.value == None:
                        app.PrintError('Celda vacía '+str(cell))
                        sys.exit()
                    else:
                        Data.append(cell.value)
        if Data!=[]:
            for i in range(len(sVar67)):
                if sVar67[i]=='idir':
                    if str(Iph.GetAttribute(sVar67[i])) == str(0) and str(Data[i]) != 'None':
                        app.PrintPlain('Diferencia en valores de Digsilent None y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='Forward':
                                Iph.SetAttribute(sVar67[i],1)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],1)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(1) and str(Data[i]) != 'Forward':
                        app.PrintPlain('Diferencia en valores de Digsilent Forward y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],2)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],2)
                    elif str(Iph.GetAttribute(sVar67[i])) == str(2) and str(Data[i]) != 'Reverse':
                        app.PrintPlain('Diferencia en valores de Digsilent Reverse y Excel '+ str(Data[i])+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            if Data[i]=='None':
                                Iph.SetAttribute(sVar67[i],0)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],0)
                            else:
                                Iph.SetAttribute(sVar67[i],1)
                                if T50==1:
                                    Iph1.SetAttribute(sVar67[i],1)
                elif sVar67[i]=='Ipset':
                    Ibase = ws.cell(row=10+iRele, column=15).value
                    Ictp = ws.cell(row=10+iRele, column=12).value
                    Icts = ws.cell(row=10+iRele, column=13).value
                    if Ibase ==  None:
                        app.PrintError('No se encontró IBase valida en fila '+str(10+iRele)+' de la columna O')
                        sys.exit()
                    if Ibase != Ictp:
                        app.PrintError('La IBase ubicada en la fila '+str(10+iRele)+' de la columna O debe ser igual a la I del CT ubicada en la fila '+str(10+iRele)+', columna L')
                        sys.exit()
                    if Unidad == 'A prim':
                        iPkup = ws.cell(row=10+iRele, column=21+T50*8).value
                        pu = iPkup/Ictp
                        iPkpn = pu
                        if abs(float(Iph.GetAttribute(sVar67[i])*lCT_p) - float(iPkpn*Ictp)) > 0.01:
                            
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])*lCT_p))+' y Excel ''{0:.2f}'.format(float(iPkpn*Ictp))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    elif Unidad == 'A sec':
                        iPkup = ws.cell(row=10+iRele, column=21+T50*8).value
                        pu = iPkup/Icts
                        iPkpn = pu
                        if abs(float(Iph.GetAttribute(sVar67[i])*Icts) - float(iPkpn*Icts)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])*Icts))+' y Excel ''{0:.2f}'.format(float(iPkpn*Icts))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    else:
                        pu = ws.cell(row=10+iRele, column=21+T50*8).value
                        iPkpn = pu
                        if abs(float(Iph.GetAttribute(sVar67[i])) - float(pu)) > 0.01:
                            app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])))+' y Excel ''{0:.2f}'.format(float(pu))+' para parametro '+str(sVar67[i])+' en el bloque '+str(Iph))
                    if Change == 'Y' or Change == 'y':
                        Iph.SetAttribute(sVar67[i],iPkpn)
                        if T50==1:
                            Iph1.SetAttribute(sVar67[i],iPkpn)
                elif sVar67[i]=='Tset':
                    if abs(float(Iph.GetAttribute(sVar67[i])) - float(Data[i])) > 0.01:
                        app.PrintPlain('Diferencia en valores de Digsilent '+'{0:.2f}'.format(float(Iph.GetAttribute(sVar67[i])))+' y Excel '+'{0:.2f}'.format(float(Data[i]))+' para parametro cptotime en el bloque '+str(Iph))
                        if Change == 'Y' or Change == 'y':
                            Iph.SetAttribute('Tset',Data[i])
                            if T50==1:
                                Iph1.SetAttribute('Tset',Data[i])

def sub_Curva_ABB(Data):
    if Data=='IEC I':
        Valor='IEC Inverse'
    elif Data=='IEC NI':
        Valor='IEC Normal inverse'
    elif Data=='IEC VI':
        Valor='IEC Very inverse'
    elif Data=='IEC EI':
        Valor='IEC Extremely inverse'
    elif Data=='IEC STI':
        Valor='IEC Short time inverse'
    elif Data=='IEC LTI':
        Valor='IEC Long time inverse'
    elif Data=='ANSI EI':
        Valor='ANSI Extremely Inverse'
    elif Data=='ANSI VI':
        Valor='ANSI Very Inverse'
    elif Data=='ANSI NI':
        Valor='ANSI Normal Inverse'
    elif Data=='ANSI MI':
        Valor='ANSI Moderately Inverse'
    elif Data=='ANSI LTEI':
        Valor='ANSI Long time Extremely Inverse'
    elif Data=='ANSI LTVI':
        Valor='ANSI Long time Very Inverse'
    elif Data=='ANSI LTI':
        Valor='ANSI Long time Inverse'
    elif Data=='RI type':
        Valor='RI inverse'
    elif Data=='RD type':
        Valor='RD Logarithmic inverse'
    return(Valor)

def sub_Curva_SIEMENS(Data):
    if Data=='IEC I':
        Valor='Definite'
    elif Data=='IEC EI':
        Valor='IEC Extremely Inverse'
    elif Data=='IEC LTI':
        Valor='IEC Long Inverse'
    elif Data=='IEC NI':
        Valor='IEC Normal Inverse'
    elif Data=='IEC VI':
        Valor='IEC Very Inverse'
    elif Data=='ANSI EI':
        Valor='ANSI Extremely Inverse'
    elif Data=='ANSI I':
        Valor='ANSI Inverse'
    elif Data=='ANSI LTI':
        Valor='ANSI Long Time Inverse'
    elif Data=='ANSI MI':
        Valor='ANSI Moderately Inverse'
    elif Data=='ANSI SI':
        Valor='ANSI Short Inverse'
    elif Data=='ANSI VI':
        Valor='ANSI Very Inverse'
    return(Valor)

def sub_Curva_SEL(Data):
    if Data=='IEC NI':
        Valor='C1 - IEC Class A (Standard Inverse)'
    elif Data=='IEC VI':
        Valor='C2 - IEC Class B (Very Inverse)'
    elif Data=='IEC EI':
        Valor='C3 - IEC Class C (Extremely Inverse)'
    elif Data=='IEC LTI':
        Valor='C4 - IEC Long Time Inverse'
    elif Data=='IEC STI':
        Valor='C5 - IEC Short Time Inverse'
    elif Data=='ANSI MI':
        Valor='U1 - U.S. Moderately Inverse'
    elif Data=='ANSI I':
        Valor='U2 - U.S. Inverse'
    elif Data=='ANSI VI':
        Valor='U3 - U.S. Very Inverse'
    elif Data=='ANSI EI':
        Valor='U4 - U.S. Extremly Inverse'
    elif Data=='ANSI SI':
        Valor='U5 - U.S. Short Time Inverse'
    return(Valor)

def sub_Curva_GE(Data):
    if Data=='ANSI EI':
        Valor='IAC Extremely inverse'
    elif Data=='ANSI NI':
        Valor='IAC Inverse'
    elif Data=='ANSI VI':
        Valor='IAC Very inverse'
    elif Data=='IEC NI':
        Valor='IEC Curve A'
    elif Data=='IEC VI':
        Valor='IEC Curve B'
    elif Data=='IEC EI':
        Valor='IEC Curve C'
    elif Data=='IEC STI':
        Valor='IEC Short Inverse'
    return(Valor)

for Case in sCs:
    app.PrintWarn('Abriendo caso de estudio '+ Case.loc_name)
    Sheets=wb.sheetnames
    for Sheet in Sheets:
        Contrl=0
        if Sheet == 'Sobrecorriente':
            Contrl=1
            ws=wb.get_sheet_by_name(Sheet)
            Cant_Reles=ws.cell(row=6, column=2).value
            if Cant_Reles == None:
                app.PrintPlain('Verifique la cantidad de relés que va añadir en la fila 6, columna B')
                sys.exit()
            iElms=[]
            sElmsLn=[]
            sElmsTer=[]
            IED_List=[]
            for iRele in range(Cant_Reles):
                sub_Crear_TM(iRele)
                sub_Crear_Reles(iRele)
    if Contrl==0:
        app.PrintPlain('La pagina donde se encuentran los ajustes de sobrecoriente en el documento de Excel debe llamarse Sobrecorriente')
        app.PrintPlain('Favor corregir y volver a ejecutar')
        sys.exit()
        
end=time.time()
app.PrintInfo('Ejecución exitosa, tiempo de ejecución de creación de relés de: {0:.2f} s'.format(end-start))


Address=script.Address
Relays=script.GetContents('Relays')[0].GetAll('ElmRelay')
Relays=list(dict.fromkeys(Relays))

def lPercent(MatDis):
    irows=MatDis.GetNumberOfRows()
    lDist=[MatDis.Get(row+1,1) for row in range(irows)]
    return lDist
lDist=lPercent(MatDis)

def Fault3F(MatRes):
    irows=MatRes.GetNumberOfRows()
    Resis3F=[MatRes.Get(row+1,1) for row in range(irows)]
    Resis3F=list(dict.fromkeys(Resis3F))
    return Resis3F
Resis3F=Fault3F(MatRes)

def Fault2F(MResistencias):
    irows=MatRes.GetNumberOfRows()
    Resis2F=[MatRes.Get(row+1,2) for row in range(irows)]
    Resis2F=list(dict.fromkeys(Resis2F))
    return Resis2F
Resis2F=Fault2F(MatRes)

def Fault1F(MResistencias):
    irows=MatRes.GetNumberOfRows()
    Resis1F=[MatRes.Get(row+1,3) for row in range(irows)]
    Resis1F=list(dict.fromkeys(Resis1F))
    return Resis1F
Resis1F=Fault1F(MatRes)

lista_resistencias=[Resis3F,Resis2F,Resis1F]

 
if Barrido == 1:
    app.PrintPlain(' ')
    app.PrintInfo('Sí se realizará el cálculo de tiempos de operación')
    
    if iBpgf not in (0,1,2):
        app.PrintInfo('Seleccione una fase fallada válida, se ejecutará la falla sobre las fases b,c')
        iBpgf=0
    if iSpgf not in (0,1,2):
        app.PrintInfo('Seleccione una fase fallada válida, se ejecutará la falla sobre la fase A')
        iSpgf=0
    
    start=time.time()
        
    ##### Obtener los tiempos de operación de cada relé del set ####
    def Tiempo_op(Relays):
        tiempo_op=[]
        for Relay in Relays:
            try:
                tiempo_op.append(Relay.GetAttribute('c:yout'))
            except:
                tiempo_op.append(str('---'))
        tiempo_op1=[t if t!=9999.999 else "N/O" for t in tiempo_op]
        return (tiempo_op1)
    
    ##### Obtener el nombre de cada relé del set ####
    def nombre_relays(Relays):
        nombre=[Relay.loc_name for Relay in Relays]
        return nombre
    
    ##### Activar los relés del set ####
    def active_relays(Relays):
        for relay in Relays:
            relay.outserv=0
    
    #### Crea un dataframe de columnas=elementos fallados ####
    def tabla_excel(Nombre,Elemento,distancia,tiempo):
        if Elemento.GetClassName()=='ElmLne' or Elemento.GetClassName()=='ElmBranch':
            dataframe=pd.DataFrame()
            dataframe['Relé']=list(Nombre)
            dataframe[Elemento.loc_name+' - '+str(distancia)+'  %']=list(tiempo)
        else:
            dataframe=pd.DataFrame()
            dataframe['Relé']=list(Nombre)
            dataframe[Elemento.loc_name]=list(tiempo)		
        return dataframe
    
    #### Une dataframes ###
    def joindataframe(tabla,Elemento,distancia,tiempo):
        if Elemento.GetClassName()=='ElmLne' or Elemento.GetClassName()=='ElmBranch':
            dataframe1=pd.DataFrame()
            dataframe1[Elemento.loc_name+' - '+str(distancia)+'  %']=tiempo
            joinedataframe=pd.concat([tabla,dataframe1],axis=1)
        else:
            dataframe1=pd.DataFrame()
            dataframe1[Elemento.loc_name]=tiempo
            joinedataframe=pd.concat([tabla,dataframe1],axis=1)
        return joinedataframe
    
    def create_writer(nombre):
        writer=pd.ExcelWriter(oFolder+'\\{}.xlsx'.format(nombre),engine='xlsxwriter')
        writer=pd.ExcelWriter(oFolder+'\\{}.xlsx'.format(nombre),engine='xlsxwriter')
        return writer
    
    ##### Exportar los dataframes a excel ####
    def export_data(writer,tabla,nombre_caso,tipo_falla,resistencia, Terminal):
        try:
            pd.formats.format.header_style=None
        except:
            pass
        workbook=writer.book
        startrow=1
        tabla=tabla.sort_values(by='Relé')
        tabla=tabla.drop_duplicates(subset='Relé')
        tabla.to_excel(writer,sheet_name=tipo_falla+" "+str(resistencia), startrow=1, index=False)
        worksheet=writer.sheets[str(tipo_falla+" "+str(resistencia))]
        worksheet.set_zoom(90)
        
        header_format=workbook.add_format({'align':'center', 'bold':True})
        header_format.set_font_name('Arial Narrow')
        header_format.set_rotation(90)
        
        parcial_fmt=workbook.add_format({'align':'center', 'num_format':'0.00'})
        parcial_fmt.set_font_name('Arial Narrow')
        
        parcial_fmt2=workbook.add_format({'align':'center','bold':1})
        parcial_fmt2.set_font_name('Arial Narrow')
        
        first_cell= workbook.add_format({'align':'center','bold': True})
        first_cell.set_font_name('Arial Narrow')
        first_cell.set_align('vcenter')
        first_cell.set_text_wrap()
        
        first_cell.set_shrink()
        texto_init='Caso de Estudio: {}\n Tipo de Falla: {}\n Resistencia: {} Ohms \n Fallas realizadas desde el terminal {}'.format(nombre_caso,tipo_falla,str(resistencia), Term)
        worksheet.write(1,0,texto_init,first_cell)
        
        worksheet.set_column('B:B',5.4,parcial_fmt)
        worksheet.set_column(2,1000,5.4,parcial_fmt)
        worksheet.set_column('A:A',60,parcial_fmt2)
        worksheet.set_row(startrow,230,header_format)
        tamaño=tabla.shape
        
        ###### Formato de cada una de las celdas
        format1=workbook.add_format({'bg_color':'#4672E0'}) ##### Tiempos elevados
        format1.set_border()
        format1.set_bold()
        format2=workbook.add_format({'bg_color':'#EC9034'}) ##### Mínimos tiempos
        format2.set_border()
        format2.set_bold()
        format3=workbook.add_format({'bg_color':'#95D7CF'}) ##### Tiempos medios
        format3.set_border()
        format3.set_bold()
        format4=workbook.add_format({'bg_color':'#FFFFFF','font_color':'#BFBFBF'}) ##### Color blanco, fuente gris
        format4.set_border()
        
        #### Formato Condicional colores ####
        worksheet.conditional_format(startrow+1,1,tamaño[0]+1,tamaño[1]-1,{'type':'cell','criteria':'equal to','value':'"N/O"','format':format4})
        worksheet.conditional_format(startrow+1,1,tamaño[0]+1,tamaño[1]-1,{'type':'cell','criteria':'<=','value':0.2001,'format':format2})
        worksheet.conditional_format(startrow+1,1,tamaño[0]+1,tamaño[1]-1,{'type':'cell','criteria':'between','minimum':0.201,'maximum':1.499,'format':format3})
        worksheet.conditional_format(startrow+1,1,tamaño[0]+1,tamaño[1]-1,{'type':'cell','criteria':'>=','value':1.5,'format':format1})
        
    ##### Clase correr cortos, extraer tiempos de Op
    class T_Op():
        def __init__(self):
            self.Metodo=iMethod
    
        @property
        def tiempo_op1(self):
            return self._tiempo_op1
        @property
        def dataframe(self):
            return self._dataframe
    	
        def Parametros_corto(self,Elemento,terminal=1,distancia=50,resistencia=0,maxima_corriente=0,iopt_shc='spgf',fFase=0):
            oShc.iopt_mde=self.Metodo
            oShc.iopt_allbus=0
            oShc.iopt_asc=0
            oShc.shcobj=Elemento
            try:
                oShc.iopt_dfr=terminal
            except:
                pass
            try:
                oShc.ppro=distancia
            except:
                pass
            oShc.iopt_shc=iopt_shc
            oShc.Rf=resistencia
            oShc.iopt_cur=maxima_corriente
            if oShc.iopt_shc=='spfg':
                oShc.i_pspgf=fFase
            if oShc.iopt_shc=='2pgf':
                oShc.i_p2pgf=fFase
            oShc.Execute()
    
            if Elemento.GetClassName()=='ElmLne' or Elemento.GetClassName()=='ElmBranch':
                app.PrintPlain('--------------------------------')
                app.PrintPlain('Corto circuito realizado en: '+ Elemento.loc_name + ' al '+ str(distancia)+ ' %')
                app.PrintPlain('Tipo de falla: '+str(iopt_shc) + ' - Rf: ' + str(resistencia)+ ' Ohms')
            else:
                app.PrintPlain('--------------------------------')
                app.PrintPlain('Corto circuito calculado en: '+ Elemento.loc_name)
                app.PrintPlain('Tipo de falla: '+str(iopt_shc) + ' - Rf: ' + str(resistencia)+ ' Ohms')
    
    #### Guarda cada uno de los objetos writers creados ###
    def save_writers(writer):
        for wr in range(len(sCs)):
            writer[wr].save()
    
    #### Determina los relés que se deben ignorar al detectar su connected branch fuera de servicio  
    def IgnoreRelays(Relays):
        IgnoreRelays=[rel for rel in Relays if rel.cbranch.IsEnergized()>0]
        return IgnoreRelays
        
    oFolder=Address+'\\{}'.format('TO Overcurrent') ### Crea folder donde almacena los documentos .xlxs ####
    if os.path.isdir(oFolder)==False:
        os.mkdir(oFolder)
    
    ldf=app.GetFromStudyCase('ComLdf')
    ldf.iopt_net=0
    
    #### Creación de los objetos Writers de los archivos de excel ####
    writer=[]
    Cases=[]
    for Case_name in sCs:
        Cases.append(Case_name.loc_name)
    for TipoF in Cases:
        writer.append(create_writer(TipoF))		
    
    ### Instancias ###
    corto_mono=T_Op()
    corto_big=T_Op()
    corto_tri=T_Op()
    
    contador=0
    for shc_case in sCs:
        shc_case.Activate()
        app.PrintPlain(' ')
        app.PrintWarn('Abriendo caso de estudio '+ shc_case.loc_name)
        nombre=nombre_relays(Relays)
        caso_name=shc_case.loc_name
        
        rx=0 ### Contador de valor de resistencia
        for resistencia in Resis1F:   
            for Element in lElms:
                ##### Líneas y Branches ####
                if Element.GetClassName()=='ElmLne' or Element.GetClassName()=='ElmBranch':
                    for distancia in lDist:
                        corto_mono.Parametros_corto(Element,terminal=Terminal,distancia=distancia,resistencia=resistencia,iopt_shc='spgf',fFase=iSpgf)      
                        tiempo=Tiempo_op(Relays)
                        
                        if Element==lElms[0] and distancia == lDist[0]:
                            Tabla_mono=tabla_excel(nombre,Element,distancia,tiempo)
                        else:
                            Tabla_mono=joindataframe(Tabla_mono,Element,distancia,tiempo)
                ##### Barras #####
                elif Element.GetClassName()=='ElmTerm':
                    distancia=0
                    corto_mono.Parametros_corto(Element,resistencia=resistencia,iopt_shc='spgf',fFase=iSpgf)
                    tiempo=Tiempo_op(Relays)
                    if Element==lElms[0]:
                        Tabla_mono=tabla_excel(nombre,Element,distancia,tiempo)
                    else:
                        Tabla_mono=joindataframe(Tabla_mono,Element,distancia,tiempo)
            ##### Almacena en memoria cada hoja de cada libro de excel ###
            export_data(writer[contador],Tabla_mono,caso_name,'spgf',resistencia, Terminal)
            rx+=1
        
        rx=0
        for resistencia in Resis2F:   
            for Element in lElms:
                ##### Líneas y Branches ####
                if Element.GetClassName()=='ElmLne' or Element.GetClassName()=='ElmBranch':
                    for distancia in lDist:
                        corto_big.Parametros_corto(Element,terminal=Terminal,distancia=distancia,resistencia=resistencia,iopt_shc='2pgf',fFase=iBpgf) 
                        tiempo=Tiempo_op(Relays)
                        if Element==lElms[0] and distancia == lDist[0]:
                            Tabla_big=tabla_excel(nombre,Element,distancia,tiempo)
                        else:
                            Tabla_big=joindataframe(Tabla_big,Element,distancia,tiempo) 
                            ##### Barras #####
                elif Element.GetClassName()=='ElmTerm':
                    distancia=0
                    corto_big.Parametros_corto(Element,resistencia=resistencia,iopt_shc='2pgf',fFase=iBpgf)
                    tiempo=Tiempo_op(Relays) 
                    if Element==lElms[0]:
                        Tabla_big=tabla_excel(nombre,Element,distancia,tiempo)
                    else:
                        Tabla_big=joindataframe(Tabla_big,Element,distancia,tiempo)
            export_data(writer[contador],Tabla_big,caso_name,'2pgf',resistencia, Terminal)
            rx+=1
        
        rx=0
        for resistencia in Resis3F:   
            for Element in lElms:
                ##### Líneas y Branches ####
                if Element.GetClassName()=='ElmLne' or Element.GetClassName()=='ElmBranch':
                    for distancia in lDist:
                        corto_tri.Parametros_corto(Element,terminal=Terminal,distancia=distancia,resistencia=resistencia,iopt_shc='3psc')
                        tiempo=Tiempo_op(Relays)
                        if Element==lElms[0] and distancia == lDist[0]:
                            Tabla_tri=tabla_excel(nombre,Element,distancia,tiempo)
                        else:
                            Tabla_tri=joindataframe(Tabla_tri,Element,distancia,tiempo)
                        ##### Barras #####
                elif Element.GetClassName()=='ElmTerm':
                    distancia=0
                    corto_tri.Parametros_corto(Element,resistencia=resistencia,iopt_shc='3psc')
                    tiempo=Tiempo_op(Relays)
                    if Element==lElms[0]:
                        Tabla_tri=tabla_excel(nombre,Element,distancia,tiempo)
                    else:
                        Tabla_tri=joindataframe(Tabla_tri,Element,distancia,tiempo)
            ##### Almacena en memoria cada hoja de cada libro de excel ###
            export_data(writer[contador],Tabla_tri,caso_name,'3psc',resistencia, Terminal)			
            rx+=1
        contador+=1
    #### Exporta los archivos almacenados en memoria a un archivo de excel
    save_writers(writer)
    end=time.time()
    
    
if Barrido == 1:
    lines=len(lElms)
    percent=len(lDist)
    for case in sCs:
        wb = load_workbook(filename=oFolder+'\\'+str(case.loc_name)+'.xlsx')
        for RF in Resis1F:
            ws = wb.get_sheet_by_name('spgf '+str(RF))
            for col in range(lines*percent):
                ws.cell(row=2, column=2+col).alignment = Alignment(textRotation=90)
                ws.freeze_panes="B3"
        for RF in Resis2F:
            ws = wb.get_sheet_by_name('2pgf '+str(RF))
            for col in range(lines*percent):
                ws.cell(row=2, column=2+col).alignment = Alignment(textRotation=90)
                ws.freeze_panes="B3"
        for RF in Resis3F:
            ws = wb.get_sheet_by_name('3psc '+str(RF))
            for col in range(lines*percent):
                ws.cell(row=2, column=2+col).alignment = Alignment(textRotation=90)
                ws.freeze_panes="B3"
            
        wb.save(oFolder+'\\'+str(case.loc_name)+'.xlsx')
        app.PrintInfo('Ejecución exitosa, tiempo de ejecución de tiempos de operación de: {0:.2f} s'.format(end-start))
        
        
if Diagnostico == 1:
    start=time.time()
    Address=script.Address
    
    GreenFill = PatternFill(start_color='66E74D', fill_type='solid')
    YellowFill = PatternFill(start_color='E2D443', fill_type='solid')
    OrangeFill = PatternFill(start_color='E7B44D', fill_type='solid')
    GrayFill = PatternFill(start_color='6F6C6C', fill_type='solid')
    RedFill = PatternFill(start_color='E25B43', fill_type='solid')
    PurpleFont = Font(name='Arial Narrow', size=11, color='AE4AF9', bold=True)
    BlackFont = Font(name='Arial Narrow', size=11, color='000000' , bold=True)
    GrayFont = Font(name='Arial Narrow', size=11, color='B2AEAD'  , bold=True)
    
    algt = Alignment(horizontal='center',vertical='center')
    algt1 = Alignment(vertical='center')
    bord = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))
    
    myList = ['Line 05 - 06','Line 05 - 06', 'Line 05 - 08','Line 08 - 09','Line 09 - 39','Line 06 - 07','Line 06 - 11','Line 07 - 08','Line 04 - 05']

    sElmsLne = list(OrderedDict.fromkeys(sElmsLn))
    
    for case in sCs:
        wb = load_workbook(filename=Address+'\\TO Overcurrent\\'+str(case.loc_name)+'.xlsx')
        Sheets_names=wb.sheetnames
        if 'Diagnostico' in Sheets_names:
            wsd = wb.get_sheet_by_name('Diagnostico')
            wb.remove_sheet(wsd)
        wsd = wb.create_sheet('Diagnostico')
        wsd.sheet_properties.tabColor="1072BA"
        
        wsd.row_dimensions[1].height = 1.0
        wsd.row_dimensions[2].height = 1.0
        wsd.column_dimensions['B'].width = 60.0
        for t in range(len(Resis1F)+len(Resis2F)+len(Resis3F)):
            wsd.row_dimensions[3+t*(len(IED_List)+10)].height = 230.0
        for t in range((len(lDist)*len(sElmsLne)+len(sElmsTer))):
            Y=get_column_letter(3+t)
            wsd.column_dimensions[Y].width = 6.0
                
        iRF=0
        for RF in Resis1F:
            pf=pd.read_excel(Address+'\\TO Overcurrent\\'+str(case.loc_name)+'.xlsx','spgf '+str(RF),header=0)
            Rel = pf.iloc[1:len(IED_List)+1,0]
            Reles = Rel.tolist()
            lLnes = pf.iloc[0,:(len(lDist)*len(sElmsLne)+len(sElmsTer)+1)]
            ilLnes = lLnes.tolist()
            for i in range((len(lDist)*len(sElmsLne)+len(sElmsTer))+1):
                if wsd.cell(row = 3+iRF*(len(IED_List)+10)   ,   column = 2+i).value==None:
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).value=ilLnes[i]
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).font  = BlackFont
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).border  = bord
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).alignment = Alignment(textRotation=90)
                Tiempos = pf.iloc[0:len(IED_List)+1,[i]]
                dFTiempos = Tiempos.iloc[0:len(IED_List)+1,0]   # para poder convertir tiempos a lista
                lTiempos = dFTiempos.tolist()                   #convierto tiempos en lista
                del lTiempos[0]                                 #elimino posición 0 de lista tiempos
                Rel_LineF=[]
                
                for b in range(len(IED_List)):
                    LFault = Tiempos.iloc[0]
                    LneFault = LFault.iloc[0]
                    if iElms[b] in LneFault:
                        Rel_LineF.append(b)
                                            
                        
                Real_Pos=[]
                for v in Rel_LineF:
                    for m in range(len(Reles)):
                        if IED_List[v]==Reles[m]:
                            Real_Pos.append(m)
                
                for j in range(len(IED_List)):
                    if wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value==None:
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value=Reles[j]
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).font  = BlackFont
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).border  = bord
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).alignment  = algt
                        
                    for k in Real_Pos:
                        if j == k:
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).value = 'N/O'
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).font  = PurpleFont
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).border = bord
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).alignment = algt
                        elif j!=k and j not in Real_Pos:
                            if lTiempos[j]!='N/O' and lTiempos[k]!='N/O':
                                if lTiempos[k]+0.18 <= lTiempos[j]:
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='OK'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = GreenFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                                elif lTiempos[k]+0.18 - lTiempos[j] <= 0.03 and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='W'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = YellowFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k]+0.18 - lTiempos[j] > 0.03 and lTiempos[k]+0.18 - lTiempos[j] < 0.18 and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='TC'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = OrangeFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k] >= lTiempos[j] and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'TC':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='D'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = RedFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                            elif lTiempos[j]=='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]!='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill  = GrayFill
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]=='N/O' and lTiempos[k]!='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).value='Relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).value='Correcta operación del dispositivo de protección respecto al relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).value='Es necesario aumentar levemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).value='Es necesario aumentar fuertemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).value='Este dispositivo de protección está operando antes que el del respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).value='Este dispositivo de protección operó, pero no operó el de la respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).value='El dispositivo de protección no presentó operación'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            '-------------------------------------'
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).font  = PurpleFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).value='OK'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).fill = GreenFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).value='W'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).fill = YellowFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).value='TC'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).fill = OrangeFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).value='D'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).fill = RedFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).fill  = GrayFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).font  = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).font  = GrayFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).alignment = algt
            
            texto_init='Caso de Estudio: {}\n Tipo de Falla: spgf\n Resistencia: {} Ohms \n Fallas realizadas desde el terminal {}'.format(case.loc_name,str(RF), Term)
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).value = texto_init
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).alignment = Alignment(horizontal='center', vertical='center',wrapText=True,textRotation=0)
            iRF+=1
            
            
        for RF in Resis2F:
            pf=pd.read_excel(Address+'\\TO Overcurrent\\'+str(case.loc_name)+'.xlsx','2pgf '+str(RF),header=0)
            Rel = pf.iloc[1:len(IED_List)+1,0]
            Reles = Rel.tolist()
            lLnes = pf.iloc[0,:(len(lDist)*len(sElmsLne)+len(sElmsTer)+1)]
            ilLnes = lLnes.tolist()
            for i in range((len(lDist)*len(sElmsLne)+len(sElmsTer))+1):
                if wsd.cell(row = 3+iRF*(len(IED_List)+10)   ,   column = 2+i).value==None:
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).value=ilLnes[i]
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).font  = BlackFont
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).border  = bord
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).alignment = Alignment(textRotation=90)
                Tiempos = pf.iloc[0:len(IED_List)+1,[i]]
                dFTiempos = Tiempos.iloc[0:len(IED_List)+1,0]   # para poder convertir tiempos a lista
                lTiempos = dFTiempos.tolist()                   #convierto tiempos en lista
                del lTiempos[0]                                 #elimino posición 0 de lista tiempos
                Rel_LineF=[]
                
                for b in range(len(IED_List)):
                    LFault = Tiempos.iloc[0]
                    LneFault = LFault.iloc[0]
                    if iElms[b] in LneFault:
                        Rel_LineF.append(b)
                        
                Real_Pos=[]
                for v in Rel_LineF:
                    for m in range(len(Reles)):
                        if IED_List[v]==Reles[m]:
                            Real_Pos.append(m)
                
                for j in range(len(IED_List)):
                    if wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value==None:
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value=Reles[j]
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).font  = BlackFont
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).border  = bord
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).alignment  = algt
                    for k in Real_Pos:
                        if j == k:
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).value = 'N/O'
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).font  = PurpleFont
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).border = bord
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).alignment = algt
                        elif j!=k and j not in Real_Pos:
                            if lTiempos[j]!='N/O' and lTiempos[k]!='N/O':
                                if lTiempos[k]+0.18 <= lTiempos[j]:
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='OK'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = GreenFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                                elif lTiempos[k]+0.18 - lTiempos[j] <= 0.03 and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='W'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = YellowFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k]+0.18 - lTiempos[j] > 0.03 and lTiempos[k]+0.18 - lTiempos[j] < 0.18 and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='TC'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = OrangeFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k] >= lTiempos[j] and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'TC':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='D'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = RedFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                            elif lTiempos[j]=='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]!='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill  = GrayFill
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]=='N/O' and lTiempos[k]!='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).value='Relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).value='Correcta operación del dispositivo de protección respecto al relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).value='Es necesario aumentar levemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).value='Es necesario aumentar fuertemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).value='Este dispositivo de protección está operando antes que el del respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).value='Este dispositivo de protección operó, pero no operó el de la respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).value='El dispositivo de protección no presentó operación'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            '-------------------------------------'
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).font  = PurpleFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).value='OK'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).fill = GreenFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).value='W'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).fill = YellowFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).value='TC'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).fill = OrangeFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).value='D'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).fill = RedFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).fill  = GrayFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).font  = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).font  = GrayFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).alignment = algt
            
            texto_init='Caso de Estudio: {}\n Tipo de Falla: 2pgf\n Resistencia: {} Ohms \n Fallas realizadas desde el terminal {}'.format(case.loc_name,str(RF), Term)
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).value = texto_init
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).alignment = Alignment(horizontal='center', vertical='center',wrapText=True,textRotation=0)
            iRF+=1
            
        for RF in Resis3F:
            pf=pd.read_excel(Address+'\\TO Overcurrent\\'+str(case.loc_name)+'.xlsx','3psc '+str(RF),header=0)
            Rel = pf.iloc[1:len(IED_List)+1,0]
            Reles = Rel.tolist()
            lLnes = pf.iloc[0,:(len(lDist)*len(sElmsLne)+len(sElmsTer)+1)]
            ilLnes = lLnes.tolist()
            for i in range((len(lDist)*len(sElmsLne)+len(sElmsTer))+1):
                if wsd.cell(row = 3+iRF*(len(IED_List)+10)   ,   column = 2+i).value==None:
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).value=ilLnes[i]
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).font  = BlackFont
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).border  = bord
                    wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2+i).alignment = Alignment(textRotation=90)
                Tiempos = pf.iloc[0:len(IED_List)+1,[i]]
                dFTiempos = Tiempos.iloc[0:len(IED_List)+1,0]   # para poder convertir tiempos a lista
                lTiempos = dFTiempos.tolist()                   #convierto tiempos en lista
                del lTiempos[0]                                 #elimino posición 0 de lista tiempos
                Rel_LineF=[]
                
                for b in range(len(IED_List)):
                    LFault = Tiempos.iloc[0]
                    LneFault = LFault.iloc[0]
                    if iElms[b] in LneFault:
                        Rel_LineF.append(b)
                        
                Real_Pos=[]
                for v in Rel_LineF:
                    for m in range(len(Reles)):
                        if IED_List[v]==Reles[m]:
                            Real_Pos.append(m)
                
                for j in range(len(IED_List)):
                    if wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value==None:
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).value=Reles[j]
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).font  = BlackFont
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).border  = bord
                        wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2).alignment  = algt
                    for k in Real_Pos:
                        if j == k:
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).value = 'N/O'
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).font  = PurpleFont
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).border = bord
                            wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,  column = 2+i).alignment = algt
                        elif j!=k and j not in Real_Pos:
                            if lTiempos[j]!='N/O' and lTiempos[k]!='N/O':
                                if lTiempos[k]+0.18 <= lTiempos[j]:
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='OK'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = GreenFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                                elif lTiempos[k]+0.18 - lTiempos[j] <= 0.03 and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='W'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = YellowFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k]+0.18 - lTiempos[j] > 0.03 and lTiempos[k]+0.18 - lTiempos[j] < 0.18 and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='TC'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = OrangeFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                                elif lTiempos[k] >= lTiempos[j] and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'OK' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'W' and wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value != 'TC':
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value='D'
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill = RedFill
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                    wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                    
                            elif lTiempos[j]=='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]!='N/O' and lTiempos[k]=='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).fill  = GrayFill
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font = BlackFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
                                
                            elif lTiempos[j]=='N/O' and lTiempos[k]!='N/O' and   wsd.cell(row=4+j+iRF*(len(IED_List)+10) , column = 2+i).value ==None:
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).value = 'N/O'
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).font  = GrayFont
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).border = bord
                                wsd.cell(row = 4+j+iRF*(len(IED_List)+10)   ,   column = 2+i).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).value='Relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+0, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).value='Correcta operación del dispositivo de protección respecto al relé perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+1, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).value='Es necesario aumentar levemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+2, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).value='Es necesario aumentar fuertemente el margen de coordinación entre este relé y el perteneciente al elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+3, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).value='Este dispositivo de protección está operando antes que el del respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+4, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).value='Este dispositivo de protección operó, pero no operó el de la respectivo elemento fallado'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+5, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).value='El dispositivo de protección no presentó operación'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 3).alignment = algt1
            wsd.merge_cells(range_string=None, start_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, start_column=3, end_row=5+len(IED_List)+iRF*(len(IED_List)+10)+6, end_column=2+(len(lDist)*len(sElmsLne)+len(sElmsTer)))
            
            '-------------------------------------'
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).font  = PurpleFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+0   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).value='OK'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).fill = GreenFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+1   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).value='W'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).fill = YellowFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+2   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).value='TC'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).fill = OrangeFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+3   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).value='D'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).fill = RedFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).font = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+4   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).fill  = GrayFill
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).font  = BlackFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+5   ,   column = 2).alignment = algt
            
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).value='N/O'
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).border = bord
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).font  = GrayFont
            wsd.cell(row = 5+len(IED_List)+iRF*(len(IED_List)+10)+6   ,   column = 2).alignment = algt
            
            texto_init='Caso de Estudio: {}\n Tipo de Falla: 3psc\n Resistencia: {} Ohms \n Fallas realizadas desde el terminal {}'.format(case.loc_name,str(RF), Term)
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).value = texto_init
            wsd.cell(row = 3+iRF*(len(IED_List)+10)  ,   column = 2).alignment = Alignment(horizontal='center', vertical='center',wrapText=True,textRotation=0)
            iRF+=1
        
        wb.save(Address+'\\TO Overcurrent\\'+str(case.loc_name)+'.xlsx')
        
    end=time.time()
    app.PrintInfo('Ejecución exitosa, tiempo de ejecución de diagnostico en: {0:.2f} s'.format(end-start))
    

app.PrintPlain("El Script ha finalizado")